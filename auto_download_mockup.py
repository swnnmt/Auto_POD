import os
import re
import time
import shutil
from pathlib import Path
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse, parse_qsl

import requests
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

BASE_DIR = Path(__file__).resolve().parent
LINK_XLSX = BASE_DIR / "LinkMockup.xlsx"
BATH_IMAGE_XLSX = BASE_DIR / "Bath_image.xlsx"
IMAGE_SRC_DIR = BASE_DIR / "Image_src"
SLEEP_AFTER_CROP = 20
MAX_LINKS = 8
PAGE_WAIT = 45
HEADLESS = False


def read_first_column_values(xlsx_path: Path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    values = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v is None:
            continue
        s = str(v).strip()
        if s:
            values.append(s)
    return values


def resolve_image_path(raw_path: str) -> Path:
    p = Path(raw_path)
    if p.is_absolute():
        return p
    candidate = (BASE_DIR / raw_path).resolve()
    if candidate.exists():
        return candidate
    candidate2 = (IMAGE_SRC_DIR / raw_path).resolve()
    if candidate2.exists():
        return candidate2
    return candidate


def build_driver():
    options = webdriver.ChromeOptions()
    if HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    prefs = {
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    }
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(options=options)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            """
        },
    )
    return driver


def try_click(wait, selectors):
    last_exc = None
    for by, value in selectors:
        try:
            el = wait.until(EC.element_to_be_clickable((by, value)))
            el.click()
            return True
        except Exception as exc:
            last_exc = exc
    if last_exc:
        raise last_exc
    return False


def upload_design_and_get_custom_param(driver, first_link: str, image_path: Path) -> str:
    wait = WebDriverWait(driver, PAGE_WAIT)
    driver.get(first_link)

    # Nhiều giao diện Placeit đổi selector liên tục, nên dùng nhiều phương án fallback.
    try_click(wait, [
        (By.XPATH, "//button[contains(., 'Design') or contains(., 'Add Design') or contains(., 'Choose Design') ]"),
        (By.XPATH, "//*[self::button or self::a or self::div][contains(., 'Design')]"),
    ])

    # Ưu tiên bắn file thẳng vào input=file nếu có.
    file_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
    sent = False
    for inp in file_inputs:
        try:
            inp.send_keys(str(image_path))
            sent = True
            break
        except Exception:
            pass

    if not sent:
        try_click(wait, [
            (By.XPATH, "//button[contains(., 'Upload From Your Device')]") ,
            (By.XPATH, "//*[self::button or self::a or self::div][contains(., 'Upload From Your Device')]")
        ])
        file_inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
        if not file_inputs:
            raise RuntimeError("Không tìm thấy input file để upload ảnh.")
        file_inputs[0].send_keys(str(image_path))

    # Crop
    try:
        try_click(wait, [
            (By.XPATH, "//button[contains(., 'Crop') or contains(., 'Apply') or contains(., 'Done') ]"),
            (By.XPATH, "//*[self::button or self::a or self::div][contains(., 'Crop')]"),
        ])
    except Exception:
        pass

    time.sleep(SLEEP_AFTER_CROP)

    current_url = driver.current_url
    parsed = urlparse(current_url)
    qs = parse_qs(parsed.query)
    custom_value = qs.get("customG_0", [None])[0]
    if not custom_value:
        m = re.search(r"[?&](customG_0=[^&]+)", current_url)
        if m:
            return m.group(1)
        raise RuntimeError(f"Không lấy được customG_0 từ URL hiện tại: {current_url}")
    return f"customG_0={custom_value}"


def append_or_replace_custom_param(url: str, custom_param: str) -> str:
    key, value = custom_param.split("=", 1)
    parsed = urlparse(url)
    q = dict(parse_qsl(parsed.query, keep_blank_values=True))
    q[key] = value
    new_query = urlencode(q)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))


def extract_best_image_url(driver, page_url: str) -> str:
    driver.get(page_url)
    WebDriverWait(driver, PAGE_WAIT).until(lambda d: d.execute_script("return document.readyState") == "complete")
    time.sleep(3)

    candidates = []

    meta_selectors = [
        "meta[property='og:image']",
        "meta[name='twitter:image']",
        "meta[property='twitter:image']",
    ]
    for sel in meta_selectors:
        els = driver.find_elements(By.CSS_SELECTOR, sel)
        for el in els:
            val = (el.get_attribute("content") or "").strip()
            if val.startswith("http"):
                candidates.append(val)

    imgs = driver.find_elements(By.CSS_SELECTOR, "img")
    for img in imgs:
        src = (img.get_attribute("src") or "").strip()
        if src.startswith("http"):
            w = img.get_attribute("naturalWidth") or img.get_attribute("width") or "0"
            h = img.get_attribute("naturalHeight") or img.get_attribute("height") or "0"
            try:
                area = int(w) * int(h)
            except Exception:
                area = 0
            if area > 200_000:
                candidates.append(src)

    seen = set()
    filtered = []
    for c in candidates:
        if c not in seen:
            seen.add(c)
            filtered.append(c)

    if not filtered:
        raise RuntimeError(f"Không tìm thấy URL ảnh trên trang: {page_url}")

    # Ưu tiên link có CDN/Placeit/mockup
    preferred = sorted(
        filtered,
        key=lambda x: (
            0 if any(k in x.lower() for k in ["placeit", "cloudfront", "mockup", "envatousercontent"]) else 1,
            len(x)
        )
    )
    return preferred[0]


def download_file(session: requests.Session, url: str, out_path: Path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with session.get(url, stream=True, timeout=120) as r:
        r.raise_for_status()
        with open(out_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=1024 * 128):
                if chunk:
                    f.write(chunk)


def copy_cookies_to_requests(driver, session: requests.Session):
    for cookie in driver.get_cookies():
        session.cookies.set(cookie["name"], cookie["value"], domain=cookie.get("domain"), path=cookie.get("path"))


def main():
    if not LINK_XLSX.exists():
        raise FileNotFoundError(f"Không thấy file: {LINK_XLSX}")
    if not BATH_IMAGE_XLSX.exists():
        raise FileNotFoundError(f"Không thấy file: {BATH_IMAGE_XLSX}")

    link_values = read_first_column_values(LINK_XLSX)
    image_values = read_first_column_values(BATH_IMAGE_XLSX)

    if not link_values:
        raise RuntimeError("LinkMockup.xlsx không có link nào ở cột A.")
    if not image_values:
        raise RuntimeError("Bath_image.xlsx không có đường dẫn ảnh nào ở cột A.")

    first_link = link_values[0]
    first_image = resolve_image_path(image_values[0])
    if not first_image.exists():
        raise FileNotFoundError(f"Không tìm thấy ảnh đầu tiên: {first_image}")

    image_stem = first_image.stem
    output_dir = IMAGE_SRC_DIR / image_stem
    output_dir.mkdir(parents=True, exist_ok=True)

    driver = build_driver()
    session = requests.Session()

    try:
        custom_param = upload_design_and_get_custom_param(driver, first_link, first_image)
        copy_cookies_to_requests(driver, session)

        target_links = link_values[:MAX_LINKS]
        final_links = [append_or_replace_custom_param(url, custom_param) for url in target_links]

        print(f"custom param lấy được: {custom_param}")
        print(f"Thư mục lưu: {output_dir}")

        for idx, link in enumerate(final_links, start=1):
            print(f"[{idx}/{len(final_links)}] Đang xử lý: {link}")
            image_url = extract_best_image_url(driver, link)
            ext = Path(urlparse(image_url).path).suffix.lower()
            if ext not in [".jpg", ".jpeg", ".png", ".webp"]:
                ext = ".jpg"
            out_file = output_dir / f"Mockup{idx}{ext}"
            download_file(session, image_url, out_file)
            print(f"Đã lưu: {out_file}")

        log_path = output_dir / "generated_links.txt"
        log_path.write_text("\n".join(final_links), encoding="utf-8")
        print("Hoàn tất.")
    finally:
        driver.quit()
        session.close()


if __name__ == "__main__":
    main()
